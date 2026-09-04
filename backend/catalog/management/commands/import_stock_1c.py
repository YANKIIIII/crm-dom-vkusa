from openpyxl import load_workbook
import re
from django.core.management.base import BaseCommand
from catalog.models import ProductCard, ProductCategory, Supplier
from warehouse.models import StockItem


class Command(BaseCommand):
    help = 'Imports stock data from 1C Excel report'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel file')
        parser.add_argument('--dry-run', action='store_true', help='Only parse and show what would be done')

    def handle(self, *args, **options):
        file_path = options['file_path']
        dry_run = options['dry_run']

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error reading file: {e}"))
            return

        products_data = self._parse(ws)
        wb.close()

        self.stdout.write(self.style.SUCCESS(f"Parsed {len(products_data)} products."))

        if dry_run:
            self.stdout.write("DRY RUN — no database changes.")
            for p in products_data:
                self.stdout.write(
                    f"  {p['name']}: {int(p['quantity'])} шт, "
                    f"себестоимость за ед.: {p['price_per_unit']:.2f} BYN"
                )
            return

        self._import(products_data)

    # ── parsing ──────────────────────────────────────────────────────

    SKIP_PREFIXES = (
        'Оборотно', 'Период', 'Детализация', 'Выводимые', 'Отбор',
        'Субконто', 'Розничный склад', 'Поступление товаров',
        'Служебный документ', 'Общество с ограниченной',
    )
    SKIP_EXACT = {'Дебет', 'Кредит', 'Итого'}

    @classmethod
    def _is_skip_row(cls, text):
        """Return True if this row is a header, sub-detail or total line."""
        if not text:
            return True
        if text in cls.SKIP_EXACT:
            return True
        return text.startswith(cls.SKIP_PREFIXES)

    @staticmethod
    def _extract_sku_and_category(name):
        name_lower = name.lower()
        sku = '-'
        clean_name = name

        # 1. Extract SKU
        # Pattern 1: Comma followed by SKU
        match = re.search(r',\s*([A-Za-zА-Яа-яЁё0-9\-]+)$', name)
        if match:
            sku = match.group(1)
            clean_name = name[:match.start()].strip()
        else:
            # Pattern 2: Last word has a dash and a digit (e.g. SA-303, MDLD316-7)
            parts = name.split()
            if parts:
                last_word = parts[-1]
                if '-' in last_word and any(c.isdigit() for c in last_word):
                    sku = last_word
                    clean_name = " ".join(parts[:-1])

        # 2. Determine Category
        if 'гриль' in name_lower and not any(kw in name_lower for kw in ['решетка', 'чехол', 'щетка', 'щипцы', 'набор', 'лопатка', 'подставка', 'сковорода', 'чугун', 'термометр']):
            cat_name, cat_code = 'Грили', 'A'
        elif any(kw in name_lower for kw in ['сковорода', 'тарелка', 'блюдо', 'миска', 'салатник', 'бокал', 'графин', 'ступка', 'форма', 'соусник', 'нож']):
            cat_name, cat_code = 'Посуда', 'DISHWARE'
        elif any(kw in name_lower for kw in ['соус', 'горчица', 'аджика', 'каперсы', 'лук', 'корнишоны', 'огурцы', 'оливки', 'перец', 'помидоры', 'кукуруза', 'заправка', 'хрен', 'соль', 'специи']):
            cat_name, cat_code = 'Специи и соусы', 'SAUCES'
        elif any(kw in name_lower for kw in ['уголь', 'розжиг', 'брикеты', 'спички', 'фен', 'стартер']):
            cat_name, cat_code = 'Расходные материалы и топливо', 'CHARCOAL'
        else:
            cat_name, cat_code = 'Аксессуары', 'B'

        grill_type = None
        if cat_code == 'A':
            if 'газовый' in name_lower:
                grill_type = 'gas'
            elif 'керамический' in name_lower:
                grill_type = 'ceramic'
            elif 'электрический' in name_lower:
                grill_type = 'electric'
            elif 'угольный' in name_lower:
                grill_type = 'charcoal'

        return clean_name, sku, cat_name, cat_code, grill_type

    def _parse(self, ws):
        """
        Parse 1C "Оборотно-сальдовая ведомость" format.

        Structure per product:
          Row A: [_, product_name, total_cost, ...]   ← we take name + cost
          Row B: [_, _,             quantity,  ...]   ← we take quantity
          Row C+: warehouse/receipt sub-rows          ← we skip these
        """
        products = []

        # Read all rows into a list (columns: A=0, B=1, C=2, ...)
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))

        i = 0
        while i < len(all_rows):
            row = all_rows[i]
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''

            if self._is_skip_row(col_b):
                i += 1
                continue

            # This should be a product name row
            product_name = col_b

            # col C (index 2) = total cost price
            try:
                total_price = float(row[2]) if row[2] is not None else 0.0
            except (ValueError, TypeError):
                i += 1
                continue

            # Next row should have quantity in col C
            if i + 1 >= len(all_rows):
                i += 1
                continue

            next_row = all_rows[i + 1]
            next_col_b = str(next_row[1]).strip() if len(next_row) > 1 and next_row[1] is not None else ''

            # The quantity row must have empty col B
            if next_col_b:
                i += 1
                continue

            try:
                quantity = float(next_row[2]) if next_row[2] is not None else 0.0
            except (ValueError, TypeError):
                i += 2
                continue

            if quantity <= 0:
                i += 2
                continue

            price_per_unit = total_price / quantity

            products.append({
                'name': product_name,
                'total_price': round(total_price, 2),
                'quantity': quantity,
                'price_per_unit': round(price_per_unit, 2),
            })

            # Skip to the row after quantity
            i += 2

        return products

    # ── database import ──────────────────────────────────────────────

    def _import(self, products_data):
        supplier, _ = Supplier.objects.get_or_create(name='Импорт 1С')

        # Cache categories
        categories_cache = {}
        for c in ProductCategory.objects.all():
            categories_cache[c.code] = c

        created = 0
        for p in products_data:
            qty = int(p['quantity'])
            cost = p['price_per_unit']
            
            clean_name, sku, cat_name, cat_code, grill_type = self._extract_sku_and_category(p['name'])
            
            if cat_code not in categories_cache:
                cat, _ = ProductCategory.objects.get_or_create(code=cat_code, defaults={'name': cat_name})
                categories_cache[cat_code] = cat
                
            category = categories_cache[cat_code]

            product, is_created = ProductCard.objects.get_or_create(
                name=clean_name,
                sku=sku,
                defaults={
                    'category': category,
                    'supplier': supplier,
                    'base_cost_price': cost,
                    'grill_type': grill_type,
                }
            )
            
            # If exists but we have new sku or category, maybe don't overwrite if it was already managed manually,
            # but since we are importing fresh, if it's created we use the extracted values.
            if is_created:
                created += 1

            StockItem.objects.update_or_create(
                product_card=product,
                defaults={'stock_quantity': qty}
            )

        self.stdout.write(self.style.SUCCESS(
            f"Import complete! Created {created} product cards with stock items."
        ))
